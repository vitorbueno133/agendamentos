import csv
import io

from flask import Blueprint, Response, jsonify, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.agendamento_service import AgendamentoService
from routes.auth_routes import token_obrigatorio

relatorio_bp = Blueprint("relatorios", __name__)
service = AgendamentoService()


def _periodo():
    """Lê início e fim da query string (?inicio=AAAA-MM-DD&fim=AAAA-MM-DD)."""
    inicio = request.args.get("inicio") or None
    fim = request.args.get("fim") or None
    return inicio, fim


@relatorio_bp.route("/api/relatorios/faturamento", methods=["GET"])
@token_obrigatorio
def faturamento():
    inicio, fim = _periodo()
    return jsonify(service.relatorio_faturamento(inicio, fim))


@relatorio_bp.route("/api/relatorios/faturamento/csv", methods=["GET"])
def faturamento_csv():
    inicio, fim = _periodo()
    dados = service.relatorio_faturamento(inicio, fim)

    saida = io.StringIO()

    # O Excel em português espera ponto-e-vírgula como separador.
    escritor = csv.writer(saida, delimiter=";")
    escritor.writerow(["Profissional", "Atendimentos", "Faturamento"])

    for linha in dados:
        valor = f"{linha['faturamento']:.2f}".replace(".", ",")

        escritor.writerow([
            linha["profissional"],
            linha["atendimentos"],
            valor,
        ])

    conteudo = "\ufeff" + saida.getvalue()

    nome = "faturamento.csv"
    if inicio or fim:
        nome = f"faturamento_{inicio or 'inicio'}_a_{fim or 'hoje'}.csv"

    return Response(
        conteudo,
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={nome}"
        },
    )


@relatorio_bp.route("/api/relatorios/faturamento/excel", methods=["GET"])
def faturamento_excel():
    inicio, fim = _periodo()
    dados = service.relatorio_faturamento(inicio, fim)

    ROXO = "7A1FA2"
    CINZA_CLARO = "F3E9F8"

    wb = Workbook()
    planilha = wb.active
    planilha.title = "Faturamento"

    if inicio or fim:
        titulo = (
            f"Relatório de Faturamento — "
            f"{inicio or 'início'} a {fim or 'hoje'}"
        )
    else:
        titulo = "Relatório de Faturamento — todo o período"

    planilha.merge_cells("A1:C1")

    celula_titulo = planilha["A1"]
    celula_titulo.value = titulo
    celula_titulo.font = Font(
        bold=True,
        size=14,
        color=ROXO,
    )
    celula_titulo.alignment = Alignment(horizontal="center")

    planilha.row_dimensions[1].height = 26

    borda = Border(
        bottom=Side(
            style="thin",
            color="CCCCCC",
        )
    )

    cabecalhos = [
        "Profissional",
        "Atendimentos",
        "Faturamento",
    ]

    for coluna, texto in enumerate(cabecalhos, start=1):
        celula = planilha.cell(
            row=3,
            column=coluna,
            value=texto,
        )
        celula.font = Font(
            bold=True,
            color="FFFFFF",
        )
        celula.fill = PatternFill(
            "solid",
            fgColor=ROXO,
        )
        celula.alignment = Alignment(horizontal="center")

    linha_atual = 4

    for indice, item in enumerate(dados):
        planilha.cell(
            row=linha_atual,
            column=1,
            value=item["profissional"],
        )

        planilha.cell(
            row=linha_atual,
            column=2,
            value=item["atendimentos"],
        )

        celula_valor = planilha.cell(
            row=linha_atual,
            column=3,
            value=item["faturamento"],
        )
        celula_valor.number_format = '"R$" #,##0.00'

        for coluna in range(1, 4):
            celula = planilha.cell(
                row=linha_atual,
                column=coluna,
            )

            celula.border = borda

            if indice % 2 == 1:
                celula.fill = PatternFill(
                    "solid",
                    fgColor=CINZA_CLARO,
                )

        planilha.cell(
            row=linha_atual,
            column=2,
        ).alignment = Alignment(horizontal="center")

        linha_atual += 1

    total_atendimentos = sum(
        item["atendimentos"] for item in dados
    )

    total_faturamento = sum(
        item["faturamento"] for item in dados
    )

    celula_total = planilha.cell(
        row=linha_atual,
        column=1,
        value="Total",
    )
    celula_total.font = Font(bold=True)

    celula_qtd = planilha.cell(
        row=linha_atual,
        column=2,
        value=total_atendimentos,
    )
    celula_qtd.font = Font(bold=True)
    celula_qtd.alignment = Alignment(horizontal="center")

    celula_soma = planilha.cell(
        row=linha_atual,
        column=3,
        value=total_faturamento,
    )
    celula_soma.font = Font(bold=True)
    celula_soma.number_format = '"R$" #,##0.00'

    for coluna in range(1, 4):
        planilha.cell(
            row=linha_atual,
            column=coluna,
        ).border = Border(
            top=Side(
                style="double",
                color=ROXO,
            )
        )

    larguras = [32, 16, 18]

    for coluna, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[
            get_column_letter(coluna)
        ].width = largura

    planilha.freeze_panes = "A4"

    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    nome = "faturamento.xlsx"

    if inicio or fim:
        nome = (
            f"faturamento_{inicio or 'inicio'}"
            f"_a_{fim or 'hoje'}.xlsx"
        )

    return Response(
        arquivo.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={nome}"
        },
    )